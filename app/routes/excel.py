import os
from io import BytesIO
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, current_app
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app import db
from app.models import Stock, Campus, User
from app.forms import UploadExcelForm, UploadEmployeesForm

excel_bp = Blueprint('excel', __name__)

EXPECTED_COLUMNS = [
    'item_name', 'category', 'quantity', 'unit', 'unit_price', 'condition',
    'status', 'asset_tag', 'serial_number', 'manufacturer', 'model', 'department',
    'assigned_to', 'remarks'
]


@excel_bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload_excel():
    form = UploadExcelForm()
    form.campus_id.choices = [(c.id, f"{c.name} ({c.code})") for c in Campus.query.order_by(Campus.name).all()]

    if form.validate_on_submit():
        file = form.file.data
        campus_id = form.campus_id.data
        campus = db.session.get(Campus, campus_id)
        if not campus:
            flash('Selected campus not found.', 'danger')
            return redirect(url_for('excel.upload_excel'))

        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            df = pd.read_excel(filepath, engine='openpyxl')
            df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]

            imported = 0
            errors = []
            for idx, row in df.iterrows():
                row_num = idx + 2  # Excel row (1-indexed header + data)
                try:
                    item_name = str(row.get('item_name', '')).strip()
                    if not item_name or item_name == 'nan':
                        errors.append(f"Row {row_num}: Missing item_name, skipped.")
                        continue

                    quantity = int(row.get('quantity', 0)) if pd.notna(row.get('quantity')) else 0
                    unit_price = float(row.get('unit_price', 0)) if pd.notna(row.get('unit_price')) else 0.0

                    # Parse optional IT asset fields
                    asset_tag = str(row.get('asset_tag', '')).strip() if pd.notna(row.get('asset_tag')) else None
                    if asset_tag == '' or asset_tag == 'nan':
                        asset_tag = None

                    # Check for duplicate asset_tag
                    if asset_tag and Stock.query.filter_by(asset_tag=asset_tag).first():
                        errors.append(f"Row {row_num}: Asset tag '{asset_tag}' already exists, skipped.")
                        continue

                    # Resolve assigned_to username
                    assigned_user_id = None
                    assigned_username = str(row.get('assigned_to', '')).strip() if pd.notna(row.get('assigned_to')) else ''
                    if assigned_username and assigned_username != 'nan':
                        user = User.query.filter(
                            db.func.lower(User.username) == assigned_username.lower()
                        ).first()
                        if user:
                            assigned_user_id = user.id
                        else:
                            errors.append(f"Row {row_num}: User '{assigned_username}' not found, asset imported unassigned.")

                    stock = Stock(
                        item_name=item_name,
                        category=str(row.get('category', '')).strip() if pd.notna(row.get('category')) else '',
                        quantity=quantity,
                        unit=str(row.get('unit', 'pcs')).strip() if pd.notna(row.get('unit')) else 'pcs',
                        unit_price=unit_price,
                        total_value=quantity * unit_price,
                        condition=str(row.get('condition', 'Good')).strip() if pd.notna(row.get('condition')) else 'Good',
                        status=str(row.get('status', 'Active')).strip() if pd.notna(row.get('status')) else 'Active',
                        asset_tag=asset_tag,
                        serial_number=str(row.get('serial_number', '')).strip() if pd.notna(row.get('serial_number')) else None,
                        manufacturer=str(row.get('manufacturer', '')).strip() if pd.notna(row.get('manufacturer')) else None,
                        model=str(row.get('model', '')).strip() if pd.notna(row.get('model')) else None,
                        department=str(row.get('department', '')).strip() if pd.notna(row.get('department')) else None,
                        assigned_to=assigned_user_id,
                        remarks=str(row.get('remarks', '')).strip() if pd.notna(row.get('remarks')) else '',
                        campus_id=campus_id,
                        added_by=current_user.username,
                    )
                    db.session.add(stock)
                    imported += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

            db.session.commit()
            flash(f'Successfully imported {imported} items to {campus.name}.', 'success')
            if errors:
                flash(f'{len(errors)} rows had issues: ' + '; '.join(errors[:5]), 'warning')

        except Exception as e:
            flash(f'Error reading Excel file: {str(e)}', 'danger')
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)

        return redirect(url_for('stock.dashboard'))

    return render_template('upload_excel.html', form=form)


@excel_bp.route('/download/template')
@login_required
def download_template():
    """Download a blank Excel template for stock data entry."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Template'

    headers = [
        'item_name', 'category', 'quantity', 'unit', 'unit_price', 'condition',
        'status', 'asset_tag', 'serial_number', 'manufacturer', 'model',
        'department', 'assigned_to', 'remarks'
    ]
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Sample rows (multiple to show variety)
    samples = [
        ['Dell Latitude 5540', 'Laptop', 1, 'pcs', 45000, 'Good',
         'Active', 'IT-2024-0001', 'SN123456', 'Dell', 'Latitude 5540', 'IT', 'john.doe', 'Assigned laptop'],
        ['HP LaserJet Pro', 'Printer', 2, 'pcs', 18000, 'Good',
         'Active', 'IT-2024-0002', 'SN789012', 'HP', 'LaserJet Pro M404', 'Admin', '', 'Floor 2 printer'],
        ['Logitech MX Keys', 'Keyboard', 5, 'pcs', 3500, 'Good',
         'In Storage', '', '', 'Logitech', 'MX Keys', 'IT', 'jane.smith', 'Wireless keyboard'],
    ]
    for row_idx, sample in enumerate(samples, 2):
        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Instructions sheet
    ws_help = wb.create_sheet(title='Instructions')
    instructions = [
        ['Column', 'Required', 'Description'],
        ['item_name', 'YES', 'Name of the asset (e.g. Dell Latitude 5540)'],
        ['category', 'No', 'Laptop, Desktop, Monitor, Printer, Scanner, Phone, Tablet, Server, Networking, etc.'],
        ['quantity', 'YES', 'Number of items (integer)'],
        ['unit', 'No', 'Unit of measure: pcs, kg, litre (default: pcs)'],
        ['unit_price', 'No', 'Price per unit (decimal number)'],
        ['condition', 'No', 'Good / Damaged / Needs Repair (default: Good)'],
        ['status', 'No', 'Active / In Storage / Under Repair / Retired / Lost-Stolen / Disposed (default: Active)'],
        ['asset_tag', 'No', 'Unique asset tag (e.g. IT-2024-0001). Must not already exist in the system.'],
        ['serial_number', 'No', 'Manufacturer serial number'],
        ['manufacturer', 'No', 'Brand / manufacturer name'],
        ['model', 'No', 'Model name or number'],
        ['department', 'No', 'Department the asset belongs to'],
        ['assigned_to', 'No', 'Username of the employee to assign this asset to. Must match an existing user in the system.'],
        ['remarks', 'No', 'Any additional notes'],
    ]
    help_header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    help_header_font = Font(bold=True, color='FFFFFF', size=11)
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_help.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = help_header_fill
                cell.font = help_header_font
    ws_help.column_dimensions['A'].width = 18
    ws_help.column_dimensions['B'].width = 10
    ws_help.column_dimensions['C'].width = 80

    # Resize data sheet columns
    wb.active = ws
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = max_length

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='stock_upload_template.xlsx'
    )


@excel_bp.route('/download/campus/<int:campus_id>')
@login_required
def download_campus_stock(campus_id):
    """Download stock data for a specific campus as Excel."""
    campus = db.session.get(Campus, campus_id)
    if not campus:
        flash('Campus not found.', 'danger')
        return redirect(url_for('stock.dashboard'))

    stocks = Stock.query.filter_by(campus_id=campus_id).order_by(Stock.category, Stock.item_name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f'{campus.code} Stock'

    # Title row
    ws.merge_cells('A1:P1')
    title_cell = ws.cell(row=1, column=1, value=f'Stock Report - {campus.name} ({campus.code})')
    title_cell.font = Font(bold=True, size=14, color='1F4E79')
    title_cell.alignment = Alignment(horizontal='center')

    # Header row
    headers = [
        'S.No', 'Asset Tag', 'Item Name', 'Category', 'Manufacturer', 'Model',
        'Serial Number', 'Quantity', 'Unit', 'Unit Price', 'Total Value',
        'Status', 'Condition', 'Assigned To', 'Department', 'Remarks'
    ]
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    grand_total = 0
    for row_idx, stock in enumerate(stocks, 4):
        sno = row_idx - 3
        total_val = (stock.quantity or 0) * (stock.unit_price or 0)
        grand_total += total_val
        assigned_name = stock.assigned_user.username if stock.assigned_user else ''

        data = [
            sno, stock.asset_tag, stock.item_name, stock.category,
            stock.manufacturer, stock.model, stock.serial_number,
            stock.quantity, stock.unit, stock.unit_price, total_val,
            stock.status, stock.condition, assigned_name, stock.department, stock.remarks
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (10, 11):  # price columns
                cell.number_format = '#,##0.00'

    # Grand total row
    total_row = len(stocks) + 4
    ws.cell(row=total_row, column=10, value='Grand Total:').font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=11, value=grand_total)
    total_cell.font = Font(bold=True, size=12)
    total_cell.number_format = '#,##0.00'

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(max_length, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'stock_{campus.code}.xlsx'
    )


@excel_bp.route('/download/all')
@login_required
def download_all_stock():
    """Download stock data for ALL campuses, each campus on a separate sheet."""
    campuses = Campus.query.order_by(Campus.name).all()
    if not campuses:
        flash('No campuses found.', 'warning')
        return redirect(url_for('stock.dashboard'))

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    headers = [
        'S.No', 'Asset Tag', 'Item Name', 'Category', 'Manufacturer', 'Model',
        'Serial Number', 'Quantity', 'Unit', 'Unit Price', 'Total Value',
        'Status', 'Condition', 'Assigned To', 'Department', 'Remarks'
    ]

    for campus in campuses:
        ws = wb.create_sheet(title=campus.code[:31])  # sheet name max 31 chars
        stocks = Stock.query.filter_by(campus_id=campus.id).order_by(Stock.category, Stock.item_name).all()

        # Title
        ws.merge_cells('A1:P1')
        title_cell = ws.cell(row=1, column=1, value=f'Stock Report - {campus.name} ({campus.code})')
        title_cell.font = Font(bold=True, size=14, color='1F4E79')
        title_cell.alignment = Alignment(horizontal='center')

        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Data
        grand_total = 0
        for row_idx, stock in enumerate(stocks, 4):
            total_val = (stock.quantity or 0) * (stock.unit_price or 0)
            grand_total += total_val
            assigned_name = stock.assigned_user.username if stock.assigned_user else ''
            data = [
                row_idx - 3, stock.asset_tag, stock.item_name, stock.category,
                stock.manufacturer, stock.model, stock.serial_number,
                stock.quantity, stock.unit, stock.unit_price, total_val,
                stock.status, stock.condition, assigned_name, stock.department, stock.remarks
            ]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in (10, 11):
                    cell.number_format = '#,##0.00'

        total_row = len(stocks) + 4
        ws.cell(row=total_row, column=10, value='Grand Total:').font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=11, value=grand_total)
        total_cell.font = Font(bold=True, size=12)
        total_cell.number_format = '#,##0.00'

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_length, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='stock_all_campuses.xlsx'
    )


# ---------------------------------------------------------------------------
# Bulk Employee Upload
# ---------------------------------------------------------------------------

@excel_bp.route('/upload/employees', methods=['GET', 'POST'])
@login_required
def upload_employees():
    if not current_user.is_admin():
        flash('Only admins can bulk-upload employees.', 'danger')
        return redirect(url_for('stock.dashboard'))

    form = UploadEmployeesForm()
    if form.validate_on_submit():
        file = form.file.data
        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            df = pd.read_excel(filepath, engine='openpyxl')
            df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]

            imported = 0
            skipped = 0
            errors = []
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    username = str(row.get('username', '')).strip()
                    if not username or username == 'nan':
                        errors.append(f"Row {row_num}: Missing username, skipped.")
                        continue

                    password = str(row.get('password', '')).strip()
                    if not password or password == 'nan':
                        password = 'changeme123'

                    # Skip if user already exists
                    if User.query.filter(db.func.lower(User.username) == username.lower()).first():
                        skipped += 1
                        errors.append(f"Row {row_num}: Username '{username}' already exists, skipped.")
                        continue

                    role = str(row.get('role', 'staff')).strip().lower() if pd.notna(row.get('role')) else 'staff'
                    if role not in ('admin', 'staff'):
                        role = 'staff'

                    department = str(row.get('department', '')).strip() if pd.notna(row.get('department')) else None
                    if department == 'nan' or department == '':
                        department = None

                    email = str(row.get('email', '')).strip() if pd.notna(row.get('email')) else None
                    if email == 'nan' or email == '':
                        email = None

                    user = User(
                        username=username,
                        role=role,
                        department=department,
                        email=email,
                    )
                    user.set_password(password)
                    db.session.add(user)
                    imported += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

            db.session.commit()
            flash(f'Successfully imported {imported} employee(s). {skipped} skipped (already exist).', 'success')
            if errors:
                flash(f'{len(errors)} row(s) had issues: ' + '; '.join(errors[:5]), 'warning')

        except Exception as e:
            flash(f'Error reading Excel file: {str(e)}', 'danger')
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)

        return redirect(url_for('stock.manage_users'))

    return render_template('upload_employees.html', form=form)


@excel_bp.route('/download/employee-template')
@login_required
def download_employee_template():
    """Download a blank Excel template for bulk employee upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employee Template'

    headers = ['username', 'password', 'role', 'department', 'email']
    header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Sample rows
    samples = [
        ['john.doe', 'password123', 'staff', 'IT', 'john.doe@company.com'],
        ['jane.smith', 'securePass!', 'staff', 'Finance', 'jane.smith@company.com'],
        ['bob.admin', 'adminPass1', 'admin', 'IT', 'bob.admin@company.com'],
    ]
    for row_idx, sample in enumerate(samples, 2):
        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Instructions sheet
    ws_help = wb.create_sheet(title='Instructions')
    instructions = [
        ['Column', 'Required', 'Description'],
        ['username', 'YES', 'Unique username for login (3-80 characters). Must not already exist in the system.'],
        ['password', 'No', 'Password for the account (min 6 chars). Defaults to "changeme123" if left blank.'],
        ['role', 'No', 'Either "admin" or "staff". Defaults to "staff" if left blank.'],
        ['department', 'No', 'Department name (e.g. IT, Finance, HR, Admin)'],
        ['email', 'No', 'Email address of the employee'],
    ]
    help_header_fill = PatternFill(start_color='1E7E34', end_color='1E7E34', fill_type='solid')
    help_header_font = Font(bold=True, color='FFFFFF', size=11)
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_help.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = help_header_fill
                cell.font = help_header_font
    ws_help.column_dimensions['A'].width = 15
    ws_help.column_dimensions['B'].width = 10
    ws_help.column_dimensions['C'].width = 80

    # Resize data sheet columns
    wb.active = ws
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = max_length

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='employee_upload_template.xlsx'
    )
